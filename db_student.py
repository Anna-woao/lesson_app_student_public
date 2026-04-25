"""学生端 Supabase 数据层（含词汇检测作答）"""

from __future__ import annotations

from io import BytesIO
import hashlib
import hmac
import random
import re
import secrets
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from supabase_client import get_admin_supabase_client, get_supabase_client


PASSWORD_HASH_ITERATIONS = 200_000
VOCAB_IMPORT_CHUNK_SIZE = 200



def _make_password_hash(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        PASSWORD_HASH_ITERATIONS,
    ).hex()
    return f"pbkdf2_sha256${PASSWORD_HASH_ITERATIONS}${salt}${digest}"


def _check_password_hash(password: str, password_hash: str) -> bool:
    try:
        algorithm, iterations, salt, expected_digest = password_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt.encode("utf-8"),
            int(iterations),
        ).hex()
    except Exception:
        return False
    return hmac.compare_digest(digest, expected_digest)


def _fetch_all_rows(query_builder, page_size: int = 1000):
    start = 0
    all_rows = []
    while True:
        resp = query_builder.range(start, start + page_size - 1).execute()
        rows = resp.data or []
        all_rows.extend(rows)
        if len(rows) < page_size:
            break
        start += page_size
    return all_rows


def _get_write_supabase_client():
    return get_admin_supabase_client() or get_supabase_client()


def _get_admin_supabase_client_required():
    client = get_admin_supabase_client()
    if client is None:
        raise RuntimeError("缺少 SUPABASE_SERVICE_ROLE_KEY，无法执行管理员写操作。")
    return client


def _normalize_lemma(text: str) -> str:
    value = (text or "").strip().lower()
    value = value.replace("’", "'").replace("‘", "'")
    value = re.sub(r"\s+", " ", value)
    return value


def _chunked(items, size: int):
    for start in range(0, len(items), size):
        yield items[start:start + size]


def _extract_pos_and_meaning(raw_meaning: str):
    text = (raw_meaning or "").strip()
    if not text:
        return "", ""

    match = re.match(r"^([A-Za-z][A-Za-z\s./-]*\.)\s*(.+)$", text)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return "", text


def get_all_students():
    supabase = get_supabase_client()
    resp = supabase.table("students").select("id, name, grade").order("id", desc=False).execute()
    rows = resp.data or []
    return [(row["id"], row["name"], row["grade"]) for row in rows]


def get_student_basic_info(student_id: int):
    supabase = get_supabase_client()
    resp = (
        supabase.table("students")
        .select("id, name, grade")
        .eq("id", student_id)
        .limit(1)
        .execute()
    )
    rows = resp.data or []
    return rows[0] if rows else None


def authenticate_student(login_account: str, login_password: str):
    account = (login_account or "").strip()
    password = (login_password or "").strip()
    if not account or not password:
        return None

    supabase = get_supabase_client()
    resp = (
        supabase.table("students")
        .select("id, name, grade, login_account, login_password, login_password_hash")
        .eq("login_account", account)
        .limit(1)
        .execute()
    )
    rows = resp.data or []
    if not rows:
        return None

    row = rows[0]
    password_hash = row.get("login_password_hash") or ""
    legacy_password = row.get("login_password") or ""

    if password_hash:
        if not _check_password_hash(password, password_hash):
            return None
    elif legacy_password:
        if not hmac.compare_digest(password, legacy_password):
            return None
        supabase.table("students").update({
            "login_password_hash": _make_password_hash(password),
            "login_password": None,
        }).eq("id", row["id"]).execute()
    else:
        return None

    return {
        "id": row["id"],
        "name": row.get("name", ""),
        "grade": row.get("grade", ""),
        "login_account": row.get("login_account", ""),
    }


def get_student_login_accounts():
    supabase = _get_admin_supabase_client_required()
    resp = (
        supabase.table("students")
        .select("id, name, grade, login_account, login_password, login_password_hash")
        .order("id", desc=False)
        .execute()
    )
    rows = resp.data or []
    return [
        (
            row["id"],
            row.get("name", ""),
            row.get("grade", ""),
            row.get("login_account", "") or "",
            bool(row.get("login_password_hash") or row.get("login_password")),
        )
        for row in rows
    ]


def update_student_login_account(student_id: int, login_account: str, login_password: str):
    account = (login_account or "").strip()
    password = (login_password or "").strip()
    if not account:
        return False, "账号不能为空。"

    supabase = _get_admin_supabase_client_required()
    duplicate_resp = (
        supabase.table("students")
        .select("id")
        .eq("login_account", account)
        .neq("id", student_id)
        .limit(1)
        .execute()
    )
    if duplicate_resp.data:
        return False, "这个账号已经被其他学生使用。"

    update_payload = {
        "login_account": account,
    }
    if password:
        update_payload["login_password_hash"] = _make_password_hash(password)
        update_payload["login_password"] = None

    supabase.table("students").update(update_payload).eq("id", student_id).execute()
    if password:
        return True, "学生账号已保存，密码已重置。"
    return True, "学生账号已保存，密码未更改。"


def parse_structured_vocab_excel(file_bytes: bytes, source_name: str = ""):
    import vocab_import_service as vis

    return vis.parse_generic_vocab_excel(
        file_bytes,
        {
            "A": "ignore",
            "B": "book_name",
            "C": "volume_name",
            "D": "unit",
            "E": "term",
            "F": "meaning",
            "G": "ipa",
            "H": "note",
        },
        source_name=source_name,
    )

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    parsed_rows = []
    errors = []

    for excel_row in range(1, worksheet.max_row + 1):
        row_values = [worksheet.cell(row=excel_row, column=col).value for col in range(1, 9)]
        raw_book_name = row_values[1]
        raw_volume_name = row_values[2]
        raw_unit = row_values[3]
        raw_surface_word = row_values[4]
        raw_meaning = row_values[5]
        raw_ipa = row_values[6]
        raw_note = row_values[7]

        book_name = str(raw_book_name or "").strip()
        volume_name = str(raw_volume_name or "").strip()
        surface_word = str(raw_surface_word or "").strip()
        meaning_text = str(raw_meaning or "").strip()
        ipa_text = str(raw_ipa or "").strip()
        note_text = str(raw_note or "").strip()

        if not any([book_name, volume_name, raw_unit, surface_word, meaning_text, ipa_text, note_text]):
            continue

        if not surface_word:
            errors.append(f"第 {excel_row} 行缺少词条/短语，已跳过。")
            continue

        if raw_unit in (None, ""):
            errors.append(f"第 {excel_row} 行缺少单元信息，已跳过。")
            continue

        unit_name = f"Unit {raw_unit}".strip()
        try:
            unit_order = int(raw_unit)
        except Exception:
            unit_order = len(parsed_rows) + 1

        pos, meaning = _extract_pos_and_meaning(meaning_text)
        normalized_lemma = _normalize_lemma(surface_word)

        parsed_rows.append({
            "row_number": excel_row,
            "book_name": book_name or "未命名词汇书",
            "volume_name": volume_name or "",
            "unit_name": unit_name,
            "unit_order": unit_order,
            "surface_word": surface_word,
            "lemma": surface_word,
            "normalized_lemma": normalized_lemma,
            "pos": pos,
            "meaning": meaning,
            "ipa": ipa_text,
            "note": note_text,
            "source_name": source_name,
        })

    return {
        "sheet_name": worksheet.title,
        "row_count": len(parsed_rows),
        "rows": parsed_rows,
        "errors": errors,
    }


def import_structured_vocab_excel(file_bytes: bytes, source_name: str = ""):
    import vocab_import_service as vis

    parsed = parse_structured_vocab_excel(file_bytes, source_name=source_name)
    return vis.import_parsed_vocab_rows(parsed, source_name=source_name)

    parsed = parse_structured_vocab_excel(file_bytes, source_name=source_name)
    rows = parsed["rows"]
    if not rows:
        return False, "没有识别到可导入的词汇行。", parsed

    supabase = _get_write_supabase_client()

    existing_books = (
        supabase.table("word_books")
        .select("id, book_name, volume_name")
        .execute()
    ).data or []
    book_map = {
        ((row.get("book_name") or "").strip(), (row.get("volume_name") or "").strip()): row["id"]
        for row in existing_books
    }

    book_ids_needed = {}
    created_book_count = 0
    for row in rows:
        key = (row["book_name"], row["volume_name"])
        book_id = book_map.get(key)
        if book_id is None:
            insert_resp = (
                supabase.table("word_books")
                .insert({
                    "book_name": row["book_name"],
                    "volume_name": row["volume_name"],
                    "description": f"由 Excel 导入：{source_name}" if source_name else "由 Excel 导入",
                })
                .execute()
            )
            inserted_rows = insert_resp.data or []
            if not inserted_rows:
                return False, f"创建词汇书失败：{row['book_name']} / {row['volume_name']}", parsed
            book_id = inserted_rows[0]["id"]
            book_map[key] = book_id
            created_book_count += 1
        book_ids_needed[key] = book_id

    existing_units = (
        supabase.table("word_units")
        .select("id, book_id, unit_name, unit_order")
        .in_("book_id", list(book_ids_needed.values()))
        .execute()
    ).data or []
    unit_map = {
        (row["book_id"], (row.get("unit_name") or "").strip()): row["id"]
        for row in existing_units
    }

    created_unit_count = 0
    for row in rows:
        book_id = book_ids_needed[(row["book_name"], row["volume_name"])]
        unit_key = (book_id, row["unit_name"])
        if unit_key in unit_map:
            continue
        insert_resp = (
            supabase.table("word_units")
            .insert({
                "book_id": book_id,
                "unit_name": row["unit_name"],
                "unit_order": row["unit_order"],
            })
            .execute()
        )
        inserted_rows = insert_resp.data or []
        if not inserted_rows:
            return False, f"创建单元失败：{row['unit_name']}", parsed
        unit_map[unit_key] = inserted_rows[0]["id"]
        created_unit_count += 1

    normalized_lemmas = []
    seen_normalized = set()
    for row in rows:
        normalized = row["normalized_lemma"]
        if not normalized or normalized in seen_normalized:
            continue
        seen_normalized.add(normalized)
        normalized_lemmas.append(normalized)

    existing_vocab_rows = []
    for chunk in _chunked(normalized_lemmas, VOCAB_IMPORT_CHUNK_SIZE):
        existing_vocab_rows.extend(
            (
                supabase.table("vocab_items")
                .select("id, lemma, normalized_lemma, pos, ipa_br, ipa_am, default_meaning")
                .in_("normalized_lemma", chunk)
                .execute()
            ).data or []
        )

    vocab_map = {
        (row.get("normalized_lemma") or "").strip(): row
        for row in existing_vocab_rows
    }

    created_vocab_count = 0
    updated_vocab_count = 0
    vocab_id_map = {}

    for row in rows:
        normalized = row["normalized_lemma"]
        existing = vocab_map.get(normalized)
        if existing:
            update_payload = {}
            if row["pos"] and row["pos"] != (existing.get("pos") or ""):
                update_payload["pos"] = row["pos"]
            if row["meaning"] and row["meaning"] != (existing.get("default_meaning") or ""):
                update_payload["default_meaning"] = row["meaning"]
            if row["ipa"]:
                if row["ipa"] != (existing.get("ipa_br") or ""):
                    update_payload["ipa_br"] = row["ipa"]
                if row["ipa"] != (existing.get("ipa_am") or ""):
                    update_payload["ipa_am"] = row["ipa"]
            if update_payload:
                supabase.table("vocab_items").update(update_payload).eq("id", existing["id"]).execute()
                existing.update(update_payload)
                updated_vocab_count += 1
            vocab_id_map[normalized] = existing["id"]
            continue

        insert_resp = (
            supabase.table("vocab_items")
            .insert({
                "lemma": row["lemma"],
                "normalized_lemma": normalized,
                "pos": row["pos"] or None,
                "ipa_br": row["ipa"] or None,
                "ipa_am": row["ipa"] or None,
                "default_meaning": row["meaning"] or None,
            })
            .execute()
        )
        inserted_rows = insert_resp.data or []
        if not inserted_rows:
            return False, f"创建词条失败：{row['lemma']}", parsed
        inserted = inserted_rows[0]
        vocab_map[normalized] = inserted
        vocab_id_map[normalized] = inserted["id"]
        created_vocab_count += 1

    target_book_ids = list(book_ids_needed.values())
    existing_links = []
    for chunk in _chunked(target_book_ids, VOCAB_IMPORT_CHUNK_SIZE):
        existing_links.extend(
            _fetch_all_rows(
                supabase.table("book_unit_vocab")
                .select("book_id, unit_id, vocab_item_id")
                .in_("book_id", chunk)
            )
        )

    existing_link_keys = {
        (row.get("book_id"), row.get("unit_id"), row.get("vocab_item_id"))
        for row in existing_links
    }

    created_link_count = 0
    for row in rows:
        book_id = book_ids_needed[(row["book_name"], row["volume_name"])]
        unit_id = unit_map[(book_id, row["unit_name"])]
        vocab_item_id = vocab_id_map[row["normalized_lemma"]]
        link_key = (book_id, unit_id, vocab_item_id)
        if link_key in existing_link_keys:
            continue
        supabase.table("book_unit_vocab").insert({
            "book_id": book_id,
            "unit_id": unit_id,
            "vocab_item_id": vocab_item_id,
            "surface_word": row["surface_word"],
            "book_meaning": row["meaning"] or None,
            "book_note": row["note"] or None,
            "source_line": f"{source_name or 'excel'}: row {row['row_number']}",
            "item_order": row["row_number"],
        }).execute()
        existing_link_keys.add(link_key)
        created_link_count += 1

    parsed["summary"] = {
        "created_books": created_book_count,
        "created_units": created_unit_count,
        "created_vocab_items": created_vocab_count,
        "updated_vocab_items": updated_vocab_count,
        "created_links": created_link_count,
        "total_rows": len(rows),
    }
    return True, "词汇书导入完成。", parsed


def get_student_recent_lessons(student_id: int, limit: int = 10):
    supabase = get_supabase_client()
    resp = (
        supabase.table("lessons")
        .select("id, lesson_type, difficulty, topic, created_at")
        .eq("student_id", student_id)
        .order("created_at", desc=True)
        .limit(limit)
        .execute()
    )
    rows = resp.data or []
    return [(r["id"], r.get("lesson_type"), r.get("difficulty"), r.get("topic"), r.get("created_at")) for r in rows]


def get_lesson_detail_for_student(student_id: int, lesson_id: int):
    supabase = get_supabase_client()
    resp = (
        supabase.table("lessons")
        .select("id, lesson_type, difficulty, topic, content, created_at")
        .eq("student_id", student_id)
        .eq("id", lesson_id)
        .limit(1)
        .execute()
    )
    rows = resp.data or []
    return rows[0] if rows else None


def get_lesson_new_vocab_for_student(student_id: int, lesson_id: int):
    supabase = get_supabase_client()
    lesson_resp = (
        supabase.table("lessons")
        .select("id")
        .eq("student_id", student_id)
        .eq("id", lesson_id)
        .limit(1)
        .execute()
    )
    if not (lesson_resp.data or []):
        return []

    link_rows = (
        supabase.table("lesson_vocab_items")
        .select("vocab_item_id, word_type")
        .eq("lesson_id", lesson_id)
        .execute()
    ).data or []

    new_vocab_ids = []
    seen_ids = set()
    for row in link_rows:
        word_type = (row.get("word_type") or "").strip().lower()
        vocab_item_id = row.get("vocab_item_id")
        if word_type not in {"new", "新词"} or vocab_item_id is None:
            continue
        if vocab_item_id in seen_ids:
            continue
        seen_ids.add(vocab_item_id)
        new_vocab_ids.append(vocab_item_id)

    if not new_vocab_ids:
        return []

    vocab_rows = (
        supabase.table("vocab_items")
        .select("id, lemma, pos, ipa_br, ipa_am, default_meaning, example_en, example_zh")
        .in_("id", new_vocab_ids)
        .execute()
    ).data or []
    vocab_map = {row["id"]: row for row in vocab_rows}

    result = []
    for vocab_item_id in new_vocab_ids:
        row = vocab_map.get(vocab_item_id)
        if not row:
            continue
        result.append({
            "id": vocab_item_id,
            "lemma": row.get("lemma", ""),
            "pos": row.get("pos", ""),
            "ipa_br": row.get("ipa_br", ""),
            "ipa_am": row.get("ipa_am", ""),
            "meaning": row.get("default_meaning", "") or "",
            "example_en": row.get("example_en", "") or "",
            "example_zh": row.get("example_zh", "") or "",
        })
    return result


def get_student_learned_vocab(student_id: int, limit: int = 200):
    supabase = get_supabase_client()
    prog = (
        supabase.table("student_vocab_progress")
        .select("vocab_item_id, status, review_count, error_count, memory_score, first_learned_at, last_review_time, next_review_time")
        .eq("student_id", student_id)
        .order("first_learned_at", desc=True)
        .limit(limit)
        .execute()
    )
    progress_rows = prog.data or []
    vocab_ids = [row["vocab_item_id"] for row in progress_rows if row.get("vocab_item_id") is not None]
    vocab_map = {}
    if vocab_ids:
        vocab_rows = (
            supabase.table("vocab_items")
            .select("id, lemma, default_meaning")
            .in_("id", vocab_ids)
            .execute()
        ).data or []
        vocab_map = {r["id"]: (r.get("lemma", ""), r.get("default_meaning", "") or "") for r in vocab_rows}

    result = []
    for row in progress_rows:
        vocab_item_id = row.get("vocab_item_id")
        lemma, meaning = vocab_map.get(vocab_item_id, ("", ""))
        if lemma:
            result.append((
                lemma,
                meaning,
                row.get("status", "learning"),
                row.get("review_count", 0),
                row.get("error_count", 0),
                row.get("memory_score", 3.0),
                row.get("first_learned_at"),
                row.get("last_review_time"),
                row.get("next_review_time"),
            ))
    return result


def get_student_learned_vocab_summary(student_id: int):
    supabase = get_supabase_client()
    progress_rows = _fetch_all_rows(
        supabase.table("student_vocab_progress")
        .select("vocab_item_id, status, review_count, error_count, memory_score, first_learned_at")
        .eq("student_id", student_id)
        .order("first_learned_at", desc=True)
    )

    progress_vocab_ids = []
    seen_progress_vocab_ids = set()
    for row in progress_rows:
        vocab_item_id = row.get("vocab_item_id")
        if vocab_item_id is None or vocab_item_id in seen_progress_vocab_ids:
            continue
        seen_progress_vocab_ids.add(vocab_item_id)
        progress_vocab_ids.append(vocab_item_id)

    if not progress_vocab_ids:
        return {"total_unique_words": 0, "lesson_groups": []}

    lesson_rows = (
        supabase.table("lessons")
        .select("id, lesson_type, topic, created_at")
        .eq("student_id", student_id)
        .order("created_at", desc=True)
        .limit(100)
        .execute()
    ).data or []

    lesson_ids = [row["id"] for row in lesson_rows if row.get("id") is not None]
    lesson_vocab_ids = {}
    linked_vocab_ids = set()

    if lesson_ids:
        link_rows = (
            supabase.table("lesson_vocab_items")
            .select("lesson_id, vocab_item_id, word_type")
            .in_("lesson_id", lesson_ids)
            .execute()
        ).data or []

        progress_vocab_id_set = set(progress_vocab_ids)
        for row in link_rows:
            lesson_id = row.get("lesson_id")
            vocab_item_id = row.get("vocab_item_id")
            if lesson_id is None or vocab_item_id is None:
                continue
            if vocab_item_id not in progress_vocab_id_set:
                continue
            lesson_vocab_ids.setdefault(lesson_id, [])
            if vocab_item_id in lesson_vocab_ids[lesson_id]:
                continue
            lesson_vocab_ids[lesson_id].append(vocab_item_id)
            linked_vocab_ids.add(vocab_item_id)

    vocab_map = {}
    if progress_vocab_ids:
        vocab_rows = (
            supabase.table("vocab_items")
            .select("id, lemma, default_meaning")
            .in_("id", progress_vocab_ids)
            .execute()
        ).data or []
        vocab_map = {
            row["id"]: {
                "lemma": row.get("lemma", ""),
                "meaning": row.get("default_meaning", "") or "",
            }
            for row in vocab_rows
        }

    lesson_groups = []
    for lesson in lesson_rows:
        lesson_id = lesson.get("id")
        vocab_list = []
        for vocab_item_id in lesson_vocab_ids.get(lesson_id, []):
            vocab = vocab_map.get(vocab_item_id)
            if not vocab or not vocab.get("lemma"):
                continue
            vocab_list.append(vocab)

        if not vocab_list:
            continue

        lesson_groups.append({
            "lesson_id": lesson_id,
            "lesson_type": lesson.get("lesson_type", ""),
            "topic": lesson.get("topic", ""),
            "created_at": lesson.get("created_at", ""),
            "word_count": len(vocab_list),
            "words": vocab_list,
        })

    ungrouped_words = []
    for vocab_item_id in progress_vocab_ids:
        if vocab_item_id in linked_vocab_ids:
            continue
        vocab = vocab_map.get(vocab_item_id)
        if not vocab or not vocab.get("lemma"):
            continue
        ungrouped_words.append(vocab)

    if ungrouped_words:
        lesson_groups.append({
            "lesson_id": None,
            "lesson_type": "未归档到学案",
            "topic": "这些词已经记入学习进度，但暂时没有关联到具体学案",
            "created_at": "",
            "word_count": len(ungrouped_words),
            "words": ungrouped_words,
        })

    return {
        "total_unique_words": len(progress_vocab_ids),
        "lesson_groups": lesson_groups,
    }


def get_student_book_progress(student_id: int):
    supabase = get_supabase_client()
    books = (
        supabase.table("word_books")
        .select("id, book_name, volume_name")
        .order("id", desc=False)
        .execute()
    ).data or []

    progress_rows = (
        supabase.table("student_vocab_progress")
        .select("vocab_item_id, status")
        .eq("student_id", student_id)
        .execute()
    ).data or []
    progress_vocab_ids = [row.get("vocab_item_id") for row in progress_rows if row.get("vocab_item_id") is not None]
    progress_vocab_map = _fetch_vocab_detail_map(progress_vocab_ids)
    progress_by_key = {}
    for row in progress_rows:
        vocab_item_id = row.get("vocab_item_id")
        progress_key = _progress_key_for_vocab(vocab_item_id, progress_vocab_map)
        if progress_key is None:
            continue
        progress_by_key[progress_key] = _merge_progress_status(
            progress_by_key.get(progress_key),
            row.get("status", "learning"),
        )

    buv_rows = _fetch_all_rows(
        supabase.table("book_unit_vocab").select("book_id, vocab_item_id")
    )
    book_vocab_ids = [row.get("vocab_item_id") for row in buv_rows if row.get("vocab_item_id") is not None]
    book_vocab_map_rows = _fetch_vocab_detail_map(book_vocab_ids)
    book_vocab_map = {}
    for row in buv_rows:
        book_id = row.get("book_id")
        vocab_item_id = row.get("vocab_item_id")
        if book_id is None or vocab_item_id is None:
            continue
        progress_key = _progress_key_for_vocab(vocab_item_id, book_vocab_map_rows)
        if progress_key is None:
            continue
        book_vocab_map.setdefault(book_id, set()).add(progress_key)

    result = []
    for book in books:
        book_id = book["id"]
        book_vocab_keys = book_vocab_map.get(book_id, set())
        learned_keys = {progress_key for progress_key in book_vocab_keys if progress_key in progress_by_key}
        mastered_count = sum(1 for progress_key in learned_keys if progress_by_key.get(progress_key) == "mastered")
        learning_count = sum(1 for progress_key in learned_keys if progress_by_key.get(progress_key) == "learning")
        review_count = sum(1 for progress_key in learned_keys if progress_by_key.get(progress_key) == "review")

        result.append((
            book_id,
            book.get("book_name"),
            book.get("volume_name"),
            len(learned_keys),
            len(book_vocab_keys),
            mastered_count,
            learning_count,
            review_count,
        ))
    return result


def get_student_unit_progress(student_id: int, book_id: int):
    supabase = get_supabase_client()
    units = (
        supabase.table("word_units")
        .select("id, unit_name, unit_order")
        .eq("book_id", book_id)
        .order("unit_order", desc=False)
        .execute()
    ).data or []

    buv_rows = _fetch_all_rows(
        supabase.table("book_unit_vocab").select("unit_id, vocab_item_id").eq("book_id", book_id)
    )
    book_vocab_ids = [row.get("vocab_item_id") for row in buv_rows if row.get("vocab_item_id") is not None]
    book_vocab_map_rows = _fetch_vocab_detail_map(book_vocab_ids)

    progress_rows = (
        supabase.table("student_vocab_progress")
        .select("vocab_item_id")
        .eq("student_id", student_id)
        .execute()
    ).data or []
    progress_vocab_ids = [row.get("vocab_item_id") for row in progress_rows if row.get("vocab_item_id") is not None]
    progress_vocab_map = _fetch_vocab_detail_map(progress_vocab_ids)
    learned_vocab_keys = {
        _progress_key_for_vocab(row.get("vocab_item_id"), progress_vocab_map)
        for row in progress_rows
        if row.get("vocab_item_id") is not None
    }
    learned_vocab_keys.discard(None)

    result = []
    for unit in units:
        unit_id = unit["id"]
        total_vocab_keys = {
            _progress_key_for_vocab(row.get("vocab_item_id"), book_vocab_map_rows)
            for row in buv_rows
            if row.get("unit_id") == unit_id and row.get("vocab_item_id") is not None
        }
        total_vocab_keys.discard(None)
        learned_in_unit = {
            progress_key
            for progress_key in total_vocab_keys
            if progress_key in learned_vocab_keys
        }
        result.append((
            unit_id,
            unit.get("unit_name"),
            unit.get("unit_order", 0),
            len(learned_in_unit),
            len(total_vocab_keys),
        ))
    return result


def get_student_vocab_test_records(student_id: int, limit: int = 20):
    supabase = get_supabase_client()
    response = (
        supabase.table("vocab_test_records")
        .select("""
            id,
            source_type,
            source_book_id,
            source_unit_id,
            source_label,
            test_type,
            test_mode,
            total_count,
            correct_count,
            accuracy,
            is_synced_to_progress,
            is_wrong_retry_round,
            created_at
        """)
        .eq("student_id", student_id)
        .order("created_at", desc=True)
        .limit(limit)
        .execute()
    )
    rows = response.data or []
    return [
        (
            row["id"],
            row.get("source_type"),
            row.get("source_book_id"),
            row.get("source_unit_id"),
            row.get("source_label"),
            row.get("test_type"),
            row.get("test_mode"),
            row.get("total_count"),
            row.get("correct_count"),
            row.get("accuracy"),
            row.get("is_synced_to_progress"),
            row.get("is_wrong_retry_round"),
            row.get("created_at"),
        )
        for row in rows
    ]


def get_student_activity_dates(student_id: int, days: int = 30):
    """
    获取某个学生近一段时间的学习活动日期。

    当前只把以下行为视为学习活动：
    - lessons.created_at：生成/保存学案
    - vocab_test_records.created_at：完成词汇检测
    """
    supabase = get_supabase_client()
    lesson_rows = (
        supabase.table("lessons")
        .select("created_at")
        .eq("student_id", student_id)
        .order("created_at", desc=True)
        .limit(max(days * 3, 30))
        .execute()
    ).data or []

    test_rows = (
        supabase.table("vocab_test_records")
        .select("created_at")
        .eq("student_id", student_id)
        .order("created_at", desc=True)
        .limit(max(days * 3, 30))
        .execute()
    ).data or []

    activity_dates = set()
    for row in lesson_rows + test_rows:
        created_at = row.get("created_at")
        if created_at:
            activity_dates.add(str(created_at)[:10])

    return sorted(activity_dates, reverse=True)[:days]


def get_latest_diagnosis_record(student_id: int):
    supabase = get_admin_supabase_client() or get_supabase_client()
    try:
        resp = (
            supabase.table("student_diagnostic_records")
            .select("*")
            .eq("student_id", student_id)
            .order("created_at", desc=True)
            .limit(1)
            .execute()
        )
    except Exception:
        return None
    rows = resp.data or []
    return rows[0] if rows else None


def get_latest_profile_snapshot(student_id: int):
    supabase = get_admin_supabase_client() or get_supabase_client()
    try:
        resp = (
            supabase.table("student_profile_snapshots")
            .select("*")
            .eq("student_id", student_id)
            .order("created_at", desc=True)
            .limit(1)
            .execute()
        )
    except Exception:
        return None
    rows = resp.data or []
    return rows[0] if rows else None


def save_initial_diagnosis_result(student_id: int, diagnosis_result: dict):
    supabase = get_admin_supabase_client() or get_supabase_client()

    record_payload = {
        "student_id": student_id,
        "diagnosis_type": "initial_mvp",
        "vocab_band": diagnosis_result.get("vocab_band"),
        "reading_profile": diagnosis_result.get("reading_profile"),
        "grammar_gap": diagnosis_result.get("grammar_gap"),
        "writing_profile": diagnosis_result.get("writing_profile"),
        "suggested_track": diagnosis_result.get("suggested_track"),
        "module_scores": diagnosis_result.get("scores", {}),
        "module_totals": diagnosis_result.get("totals", {}),
    }
    record_resp = supabase.table("student_diagnostic_records").insert(record_payload).execute()
    record_rows = record_resp.data or []
    record = record_rows[0] if record_rows else None
    record_id = record.get("id") if record else None

    snapshot_payload = {
        "student_id": student_id,
        "source_record_id": record_id,
        "source_type": "initial_diagnosis",
        "title_label": diagnosis_result.get("title_label"),
        "stage_label": diagnosis_result.get("stage_label"),
        "growth_focus": diagnosis_result.get("growth_focus"),
        "summary_text": diagnosis_result.get("summary_text"),
        "profile_payload": {
            "dimensions": diagnosis_result.get("dimensions", {}),
            "suggested_track": diagnosis_result.get("suggested_track"),
            "vocab_band": diagnosis_result.get("vocab_band"),
            "reading_profile": diagnosis_result.get("reading_profile"),
            "grammar_gap": diagnosis_result.get("grammar_gap"),
            "writing_profile": diagnosis_result.get("writing_profile"),
        },
    }
    snapshot_resp = supabase.table("student_profile_snapshots").insert(snapshot_payload).execute()
    snapshot_rows = snapshot_resp.data or []
    snapshot = snapshot_rows[0] if snapshot_rows else None

    return {
        "record": record,
        "snapshot": snapshot,
    }


def get_vocab_test_record_items(test_record_id: int):
    supabase = get_supabase_client()
    response = (
        supabase.table("vocab_test_record_items")
        .select("vocab_item_id, word, meaning, mode, user_answer, is_correct")
        .eq("test_record_id", test_record_id)
        .order("id", desc=False)
        .execute()
    )
    rows = response.data or []
    return [
        (
            row.get("vocab_item_id"),
            row.get("word"),
            row.get("meaning"),
            row.get("mode"),
            row.get("user_answer"),
            row.get("is_correct"),
        )
        for row in rows
    ]


def get_all_word_books():
    supabase = get_supabase_client()
    rows = (
        supabase.table("word_books")
        .select("id, book_name, volume_name")
        .order("id", desc=False)
        .execute()
    ).data or []
    result = []
    for row in rows:
        label = row["book_name"] if not row.get("volume_name") else f"{row['book_name']}（{row['volume_name']}）"
        result.append((row["id"], label))
    return result


def get_units_by_book(book_id: int):
    supabase = get_supabase_client()
    rows = (
        supabase.table("word_units")
        .select("id, unit_name, unit_order")
        .eq("book_id", book_id)
        .order("unit_order", desc=False)
        .execute()
    ).data or []
    return [(r["id"], r["unit_name"], r.get("unit_order", 0)) for r in rows]


def _fetch_vocab_map(vocab_ids: List[int]) -> Dict[int, Tuple[str, str]]:
    supabase = get_supabase_client()
    if not vocab_ids:
        return {}
    rows = (
        supabase.table("vocab_items")
        .select("id, lemma, default_meaning")
        .in_("id", vocab_ids)
        .execute()
    ).data or []
    return {r["id"]: (r.get("lemma", ""), r.get("default_meaning", "") or "") for r in rows}


def _fetch_vocab_detail_map(vocab_ids: List[int]) -> Dict[int, dict]:
    supabase = _get_admin_supabase_client_required()
    if not vocab_ids:
        return {}

    rows = []
    deduped_ids = list(dict.fromkeys(vocab_ids))
    batch_size = 500
    for start in range(0, len(deduped_ids), batch_size):
        batch_ids = deduped_ids[start:start + batch_size]
        try:
            batch_rows = (
                supabase.table("vocab_items")
                .select("id, lemma, normalized_lemma, default_meaning")
                .in_("id", batch_ids)
                .execute()
            ).data or []
        except Exception as exc:
            error_text = str(exc)
            if "normalized_lemma" not in error_text and "PGRST204" not in error_text:
                raise
            batch_rows = (
                supabase.table("vocab_items")
                .select("id, lemma, default_meaning")
                .in_("id", batch_ids)
                .execute()
            ).data or []
        rows.extend(batch_rows)

    for row in rows:
        row["normalized_lemma"] = _normalize_lemma(row.get("normalized_lemma") or row.get("lemma") or "")
    return {r["id"]: r for r in rows}


def _progress_key_for_vocab(vocab_item_id: Optional[int], vocab_rows_map: Dict[int, dict]) -> Optional[str]:
    if vocab_item_id is None:
        return None
    vocab_row = vocab_rows_map.get(vocab_item_id, {})
    normalized = (vocab_row.get("normalized_lemma") or "").strip()
    return normalized or f"id:{vocab_item_id}"


def _merge_progress_status(current_status: Optional[str], new_status: Optional[str]) -> str:
    order = {"learning": 1, "review": 2, "mastered": 3}
    current = current_status or "learning"
    incoming = new_status or "learning"
    return incoming if order.get(incoming, 0) >= order.get(current, 0) else current


def build_progress_test(student_id: int, test_type: str, test_mode: str, test_count: int):
    supabase = get_supabase_client()
    status_filter = "learning" if test_type == "新词检测" else "review"
    rows = (
        supabase.table("student_vocab_progress")
        .select("vocab_item_id, first_source_book_id, first_source_unit_id, status")
        .eq("student_id", student_id)
        .eq("status", status_filter)
        .limit(500)
        .execute()
    ).data or []
    if not rows:
        return False, "当前没有可用的学习进度检测词。"
    random.shuffle(rows)
    rows = rows[:test_count]

    vocab_map = _fetch_vocab_map([r["vocab_item_id"] for r in rows if r.get("vocab_item_id") is not None])
    questions = _build_questions_from_vocab_map(vocab_map, rows, test_mode)
    payload = {
        "source_type": "progress",
        "source_book_id": None,
        "source_unit_id": None,
        "test_type": test_type,
        "test_mode": test_mode,
        "questions": questions,
    }
    return True, payload


def build_book_test(
    student_id: int,
    book_id: int,
    unit_ids: Optional[List[int]],
    test_mode: str,
    test_count: int,
):
    """
    从词汇书中抽题，支持“整本书”或“多个单元联合检测”。

    参数说明：
    - student_id:
        当前学生 ID（这版先保留，后面如果要做更复杂的个性化抽词还可以继续用）
    - book_id:
        当前选择的词汇书 ID
    - unit_ids:
        None 或 []   -> 表示整本词汇书
        [1, 2, 3]    -> 表示多个单元联合检测
    - test_mode:
        英译中 / 中译英 / 混合模式
    - test_count:
        本轮检测题数
    """
    supabase = get_supabase_client()

    query = supabase.table("book_unit_vocab").select("vocab_item_id, book_id, unit_id")
    query = query.eq("book_id", book_id)

    # ------------------------------
    # 如果传了多个单元，就只从这些单元里抽
    # 如果为空，就默认整本词汇书
    # ------------------------------
    if unit_ids:
        query = query.in_("unit_id", unit_ids)

    rows = _fetch_all_rows(query)
    if not rows:
        return False, "当前范围内没有可用单词。"

    # ------------------------------
    # 去重：
    # 1. 避免同一个词因为出现在多个单元里被重复抽到
    # 2. 这样联合多单元检测时更稳定
    # ------------------------------
    deduped_rows = []
    seen_vocab_ids = set()

    for row in rows:
        vocab_item_id = row.get("vocab_item_id")
        if vocab_item_id is None:
            continue
        if vocab_item_id in seen_vocab_ids:
            continue

        seen_vocab_ids.add(vocab_item_id)
        deduped_rows.append(row)

    random.shuffle(deduped_rows)
    deduped_rows = deduped_rows[:test_count]

    vocab_map = _fetch_vocab_map(
        [row["vocab_item_id"] for row in deduped_rows if row.get("vocab_item_id") is not None]
    )
    questions = _build_questions_from_vocab_map(vocab_map, deduped_rows, test_mode)

    payload = {
        "source_type": "book",
        "source_book_id": book_id,
        # 数据表里目前只有一个 source_unit_id 字段
        # 所以：
        # - 只选了一个单元：正常写这个 unit_id
        # - 选了多个单元 / 整本书：这里先记 None
        "source_unit_id": unit_ids[0] if unit_ids and len(unit_ids) == 1 else None,
        "selected_unit_ids": unit_ids or [],
        "test_type": "词汇书抽词检测",
        "test_mode": test_mode,
        "questions": questions,
    }
    return True, payload


def _build_questions_from_vocab_map(vocab_map, rows, test_mode: str):
    all_meanings = [meaning for _, meaning in vocab_map.values() if meaning]
    questions = []

    for row in rows:
        vocab_item_id = row.get("vocab_item_id")
        word, meaning = vocab_map.get(vocab_item_id, ("", ""))
        if not word:
            continue

        one_mode = random.choice(["英译中", "中译英"]) if test_mode == "混合模式" else test_mode

        q = {
            "vocab_item_id": vocab_item_id,
            "word": word,
            "meaning": meaning,
            "mode": one_mode,
        }

        if one_mode == "英译中":
            distractors = [m for m in all_meanings if m and m != meaning]
            random.shuffle(distractors)
            options = [meaning] + distractors[:3]
            random.shuffle(options)
            q["options"] = options

        questions.append(q)

    return questions


def submit_student_test(student_id: int, payload: dict, user_answers: dict, source_label: str):
    results = []
    score = 0
    total = len(payload["questions"])

    for q in payload["questions"]:
        vocab_item_id = q["vocab_item_id"]
        mode = q["mode"]
        user_answer = user_answers.get(vocab_item_id, "")
        if mode == "英译中":
            is_correct = (user_answer or "").strip() == (q["meaning"] or "").strip()
        else:
            is_correct = (user_answer or "").strip().lower() == (q["word"] or "").strip().lower()

        if is_correct:
            score += 1

        results.append({
            "vocab_item_id": vocab_item_id,
            "word": q["word"],
            "meaning": q["meaning"],
            "mode": mode,
            "user_answer": user_answer,
            "is_correct": is_correct,
        })

    accuracy = (score / total) if total else 0.0

    test_record_id = None
    persistence_error = None
    try:
        supabase = get_supabase_client()
        record_resp = (
            supabase.table("vocab_test_records")
            .insert({
                "student_id": student_id,
                "source_type": payload.get("source_type"),
                "source_book_id": payload.get("source_book_id"),
                "source_unit_id": payload.get("source_unit_id"),
                "source_label": source_label,
                "test_type": payload.get("test_type"),
                "test_mode": payload.get("test_mode"),
                "total_count": total,
                "correct_count": score,
                "accuracy": accuracy,
                "is_synced_to_progress": False,
                "is_wrong_retry_round": False,
            })
            .execute()
        )
        record_rows = record_resp.data or []
        if record_rows:
            test_record_id = record_rows[0]["id"]

        if test_record_id:
            item_payloads = []
            for r in results:
                item_payloads.append({
                    "test_record_id": test_record_id,
                    "vocab_item_id": r["vocab_item_id"],
                    "word": r["word"],
                    "meaning": r["meaning"],
                    "mode": r["mode"],
                    "user_answer": r["user_answer"],
                    "is_correct": r["is_correct"],
                })
            if item_payloads:
                supabase.table("vocab_test_record_items").insert(item_payloads).execute()
    except Exception as exc:
        persistence_error = str(exc)

    return {
        "score": score,
        "total": total,
        "accuracy": accuracy,
        "results": results,
        "test_record_id": test_record_id,
        "persistence_ok": persistence_error is None,
        "persistence_error": persistence_error,
    }
