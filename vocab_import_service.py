"""Generic Excel vocab import with reusable templates."""

from __future__ import annotations

from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook

from supabase_client import get_admin_supabase_client, get_supabase_client


VOCAB_IMPORT_CHUNK_SIZE = 200
COLUMN_ROLES = [
    "ignore",
    "book_name",
    "volume_name",
    "unit",
    "term",
    "meaning",
    "ipa",
    "pos",
    "note",
]


def _get_write_supabase_client():
    client = get_admin_supabase_client()
    if client is None:
        raise RuntimeError("缺少 SUPABASE_SERVICE_ROLE_KEY，无法执行词汇导入。")
    return client


def _get_template_store_client():
    client = get_admin_supabase_client()
    if client is None:
        raise RuntimeError("缺少 SUPABASE_SERVICE_ROLE_KEY，无法读取或保存导入模板。")
    return client


def _fetch_all_rows(query_builder, page_size: int = 1000):
    start = 0
    all_rows = []
    while True:
        resp = query_builder.range(start, start + page_size - 1).execute()
        rows = resp.data or []
        all_rows.extend(rows)
        if len(rows) < page_size:
            break
        start += page_size
    return all_rows


def _chunked(items, size: int):
    for start in range(0, len(items), size):
        yield items[start:start + size]


def _normalize_lemma(text: str) -> str:
    value = (text or "").strip().lower()
    value = value.replace("’", "'").replace("‘", "'")
    value = re.sub(r"\s+", " ", value)
    return value


def _extract_pos_and_meaning(raw_meaning: str, explicit_pos: str = ""):
    pos = (explicit_pos or "").strip()
    text = (raw_meaning or "").strip()
    if pos:
        return pos, text
    if not text:
        return "", ""
    match = re.match(r"^([A-Za-z][A-Za-z\s./-]*\.)\s*(.+)$", text)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return "", text


def _looks_like_ipa(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if "/" in text or "[" in text or "]" in text:
        return True
    ipa_chars = "ˈˌəɪʊæɑɒɔʃʒθðŋ"
    return any(ch in text for ch in ipa_chars)


def _looks_like_meaning(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if re.search(r"[\u4e00-\u9fff]", text):
        return True
    return any(text.startswith(prefix) for prefix in ["n.", "v.", "adj.", "adv.", "prep.", "conj.", "pron."])


def _looks_like_term(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if re.search(r"[\u4e00-\u9fff]", text):
        return False
    return bool(re.search(r"[A-Za-z]", text))


def _looks_like_book_name(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    keywords = ["版", "外研", "人教", "译林", "北师", "牛津", "教材"]
    return any(key in text for key in keywords)


def _looks_like_volume_name(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    keywords = ["必修", "选择性必修", "册", "上册", "下册"]
    return any(key in text for key in keywords)


def _looks_like_unit(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    if re.fullmatch(r"\d+", text):
        return True
    return text.lower().startswith("unit ")


def _is_serial_column(samples: list[str]) -> bool:
    numbers = []
    for value in samples:
        text = str(value).strip()
        if not re.fullmatch(r"\d+", text):
            return False
        numbers.append(int(text))
    if len(numbers) < 3:
        return False
    return numbers == list(range(numbers[0], numbers[0] + len(numbers)))


def _excel_col_label(index: int) -> str:
    label = ""
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        label = chr(65 + remainder) + label
    return label


def _read_worksheet(file_bytes: bytes, sheet_name: str | None = None):
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    return workbook.sheetnames, worksheet.title, rows


def _safe_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def preview_excel(file_bytes: bytes, sheet_name: str | None = None, preview_rows: int = 20):
    sheet_names, active_sheet, rows = _read_worksheet(file_bytes, sheet_name=sheet_name)
    max_cols = max((len(row) for row in rows), default=0)
    headers = [_excel_col_label(idx) for idx in range(1, max_cols + 1)]
    preview = []
    for row in rows[:preview_rows]:
        normalized = [_safe_cell(row[idx]) if idx < len(row) else "" for idx in range(max_cols)]
        preview.append(normalized)
    return {
        "sheet_names": sheet_names,
        "active_sheet": active_sheet,
        "headers": headers,
        "preview_rows": preview,
        "row_count": len(rows),
        "column_count": max_cols,
    }


def guess_mapping(file_bytes: bytes, sheet_name: str | None = None, sample_rows: int = 30):
    preview = preview_excel(file_bytes, sheet_name=sheet_name, preview_rows=sample_rows)
    headers = preview["headers"]
    rows = preview["preview_rows"]
    mapping = {header: "ignore" for header in headers}

    role_best = {role: (None, -1) for role in ["book_name", "volume_name", "unit", "term", "meaning", "ipa"]}
    for idx, header in enumerate(headers):
        samples = [row[idx] for row in rows if idx < len(row) and row[idx]]
        if not samples:
            continue
        unique_count = len(set(samples))
        unique_ratio = unique_count / max(len(samples), 1)
        is_serial = _is_serial_column(samples)
        scores = {
            "book_name": sum(2 for value in samples if _looks_like_book_name(value)) + (5 if unique_ratio <= 0.3 else 0),
            "volume_name": sum(2 for value in samples if _looks_like_volume_name(value)) + (5 if unique_ratio <= 0.3 else 0),
            "unit": (0 if is_serial else sum(2 for value in samples if _looks_like_unit(value))) + (3 if not is_serial and unique_ratio <= 0.4 else 0),
            "term": sum(2 for value in samples if _looks_like_term(value)) + (4 if unique_ratio >= 0.6 else 0),
            "meaning": sum(2 for value in samples if _looks_like_meaning(value)) + (4 if unique_ratio >= 0.6 else 0),
            "ipa": sum(1 for value in samples if _looks_like_ipa(value)),
        }
        for role, score in scores.items():
            if score > role_best[role][1]:
                role_best[role] = (header, score)

    used_headers = set()
    for role, (header, score) in role_best.items():
        if header and score > 0 and header not in used_headers:
            mapping[header] = role
            used_headers.add(header)

    return mapping


def _load_templates():
    client = _get_template_store_client()
    try:
        rows = (
            client.table("vocab_import_templates")
            .select("name, mapping, sheet_name, data_start_row, updated_at")
            .order("name", desc=False)
            .execute()
        ).data or []
    except Exception:
        return []
    return rows


def list_import_templates():
    return _load_templates()


def save_import_template(name: str, template: dict):
    client = _get_template_store_client()
    payload = {
        "name": name,
        "mapping": template.get("mapping", {}),
        "sheet_name": template.get("sheet_name"),
        "data_start_row": int(template.get("data_start_row", 1) or 1),
    }
    client.table("vocab_import_templates").upsert(payload, on_conflict="name").execute()


def delete_import_template(name: str):
    client = _get_template_store_client()
    client.table("vocab_import_templates").delete().eq("name", name).execute()


def _get_template_by_name(name: str):
    for item in _load_templates():
        if item.get("name") == name:
            return item
    return None


def parse_generic_vocab_excel(
    file_bytes: bytes,
    mapping: dict[str, str],
    *,
    sheet_name: str | None = None,
    data_start_row: int = 1,
    source_name: str = "",
):
    _sheet_names, active_sheet, rows = _read_worksheet(file_bytes, sheet_name=sheet_name)
    max_cols = max((len(row) for row in rows), default=0)
    headers = [_excel_col_label(idx) for idx in range(1, max_cols + 1)]

    column_by_role = {}
    for header in headers:
        role = mapping.get(header, "ignore")
        if role and role != "ignore":
            column_by_role[role] = headers.index(header)

    if "term" not in column_by_role:
        return {
            "sheet_name": active_sheet,
            "row_count": 0,
            "rows": [],
            "errors": ["必须指定词条列。"],
        }

    if "unit" not in column_by_role:
        return {
            "sheet_name": active_sheet,
            "row_count": 0,
            "rows": [],
            "errors": ["必须指定单元列。"],
        }

    parsed_rows = []
    errors = []

    for excel_row in range(data_start_row, len(rows) + 1):
        row = rows[excel_row - 1]

        def get_role(role: str):
            col_idx = column_by_role.get(role)
            if col_idx is None or col_idx >= len(row):
                return ""
            return _safe_cell(row[col_idx])

        book_name = get_role("book_name")
        volume_name = get_role("volume_name")
        raw_unit = get_role("unit")
        surface_word = get_role("term")
        meaning_text = get_role("meaning")
        ipa_text = get_role("ipa")
        explicit_pos = get_role("pos")
        note_text = get_role("note")

        if not any([book_name, volume_name, raw_unit, surface_word, meaning_text, ipa_text, explicit_pos, note_text]):
            continue
        if not surface_word:
            errors.append(f"第 {excel_row} 行缺少词条/短语，已跳过。")
            continue
        if not raw_unit:
            errors.append(f"第 {excel_row} 行缺少单元信息，已跳过。")
            continue

        unit_text = str(raw_unit).strip()
        unit_name = unit_text if unit_text.lower().startswith("unit ") else f"Unit {unit_text}"
        unit_match = re.search(r"(\d+)", unit_text)
        unit_order = int(unit_match.group(1)) if unit_match else len(parsed_rows) + 1
        pos, meaning = _extract_pos_and_meaning(meaning_text, explicit_pos=explicit_pos)

        parsed_rows.append({
            "row_number": excel_row,
            "book_name": book_name or "未命名词汇书",
            "volume_name": volume_name or "",
            "unit_name": unit_name,
            "unit_order": unit_order,
            "surface_word": surface_word,
            "lemma": surface_word,
            "normalized_lemma": _normalize_lemma(surface_word),
            "pos": pos,
            "meaning": meaning,
            "ipa": ipa_text,
            "note": note_text,
            "source_name": source_name,
        })

    return {
        "sheet_name": active_sheet,
        "row_count": len(parsed_rows),
        "rows": parsed_rows,
        "errors": errors,
    }


def parse_with_template(file_bytes: bytes, template_name: str, source_name: str = ""):
    template = _get_template_by_name(template_name)
    if not template:
        return {
            "sheet_name": "",
            "row_count": 0,
            "rows": [],
            "errors": [f"没有找到模板：{template_name}"],
        }
    return parse_generic_vocab_excel(
        file_bytes,
        template.get("mapping", {}),
        sheet_name=template.get("sheet_name") or None,
        data_start_row=int(template.get("data_start_row", 1) or 1),
        source_name=source_name,
    )


def import_parsed_vocab_rows(parsed: dict, source_name: str = ""):
    rows = parsed["rows"]
    if not rows:
        return False, "没有识别到可导入的词汇行。", parsed

    supabase = _get_write_supabase_client()

    existing_books = (
        supabase.table("word_books")
        .select("id, book_name, volume_name")
        .execute()
    ).data or []
    book_map = {
        ((row.get("book_name") or "").strip(), (row.get("volume_name") or "").strip()): row["id"]
        for row in existing_books
    }

    book_ids_needed = {}
    created_book_count = 0
    for row in rows:
        key = (row["book_name"], row["volume_name"])
        book_id = book_map.get(key)
        if book_id is None:
            insert_resp = (
                supabase.table("word_books")
                .insert({
                    "book_name": row["book_name"],
                    "volume_name": row["volume_name"],
                    "description": f"Imported from Excel: {source_name}" if source_name else "Imported from Excel",
                })
                .execute()
            )
            inserted_rows = insert_resp.data or []
            if not inserted_rows:
                return False, f"创建词汇书失败：{row['book_name']} / {row['volume_name']}", parsed
            book_id = inserted_rows[0]["id"]
            book_map[key] = book_id
            created_book_count += 1
        book_ids_needed[key] = book_id

    existing_units = (
        supabase.table("word_units")
        .select("id, book_id, unit_name, unit_order")
        .in_("book_id", list(book_ids_needed.values()))
        .execute()
    ).data or []
    unit_map = {
        (row["book_id"], (row.get("unit_name") or "").strip()): row["id"]
        for row in existing_units
    }

    created_unit_count = 0
    for row in rows:
        book_id = book_ids_needed[(row["book_name"], row["volume_name"])]
        unit_key = (book_id, row["unit_name"])
        if unit_key in unit_map:
            continue
        insert_resp = (
            supabase.table("word_units")
            .insert({
                "book_id": book_id,
                "unit_name": row["unit_name"],
                "unit_order": row["unit_order"],
            })
            .execute()
        )
        inserted_rows = insert_resp.data or []
        if not inserted_rows:
            return False, f"创建单元失败：{row['unit_name']}", parsed
        unit_map[unit_key] = inserted_rows[0]["id"]
        created_unit_count += 1

    normalized_lemmas = []
    seen_normalized = set()
    for row in rows:
        normalized = row["normalized_lemma"]
        if not normalized or normalized in seen_normalized:
            continue
        seen_normalized.add(normalized)
        normalized_lemmas.append(normalized)

    existing_vocab_rows = []
    for chunk in _chunked(normalized_lemmas, VOCAB_IMPORT_CHUNK_SIZE):
        existing_vocab_rows.extend(
            (
                supabase.table("vocab_items")
                .select("id, lemma, normalized_lemma, pos, ipa_br, ipa_am, default_meaning")
                .in_("normalized_lemma", chunk)
                .execute()
            ).data or []
        )

    vocab_map = {(row.get("normalized_lemma") or "").strip(): row for row in existing_vocab_rows}
    created_vocab_count = 0
    updated_vocab_count = 0
    vocab_id_map = {}

    for row in rows:
        normalized = row["normalized_lemma"]
        existing = vocab_map.get(normalized)
        if existing:
            update_payload = {}
            if row["pos"] and row["pos"] != (existing.get("pos") or ""):
                update_payload["pos"] = row["pos"]
            if row["meaning"] and row["meaning"] != (existing.get("default_meaning") or ""):
                update_payload["default_meaning"] = row["meaning"]
            if row["ipa"]:
                if row["ipa"] != (existing.get("ipa_br") or ""):
                    update_payload["ipa_br"] = row["ipa"]
                if row["ipa"] != (existing.get("ipa_am") or ""):
                    update_payload["ipa_am"] = row["ipa"]
            if update_payload:
                supabase.table("vocab_items").update(update_payload).eq("id", existing["id"]).execute()
                existing.update(update_payload)
                updated_vocab_count += 1
            vocab_id_map[normalized] = existing["id"]
            continue

        insert_resp = (
            supabase.table("vocab_items")
            .insert({
                "lemma": row["lemma"],
                "normalized_lemma": normalized,
                "pos": row["pos"] or None,
                "ipa_br": row["ipa"] or None,
                "ipa_am": row["ipa"] or None,
                "default_meaning": row["meaning"] or None,
            })
            .execute()
        )
        inserted_rows = insert_resp.data or []
        if not inserted_rows:
            return False, f"创建词条失败：{row['lemma']}", parsed
        inserted = inserted_rows[0]
        vocab_map[normalized] = inserted
        vocab_id_map[normalized] = inserted["id"]
        created_vocab_count += 1

    target_book_ids = list(book_ids_needed.values())
    existing_links = []
    for chunk in _chunked(target_book_ids, VOCAB_IMPORT_CHUNK_SIZE):
        existing_links.extend(
            _fetch_all_rows(
                supabase.table("book_unit_vocab")
                .select("book_id, unit_id, vocab_item_id")
                .in_("book_id", chunk)
            )
        )

    existing_link_keys = {
        (row.get("book_id"), row.get("unit_id"), row.get("vocab_item_id"))
        for row in existing_links
    }

    created_link_count = 0
    for row in rows:
        book_id = book_ids_needed[(row["book_name"], row["volume_name"])]
        unit_id = unit_map[(book_id, row["unit_name"])]
        vocab_item_id = vocab_id_map[row["normalized_lemma"]]
        link_key = (book_id, unit_id, vocab_item_id)
        if link_key in existing_link_keys:
            continue
        supabase.table("book_unit_vocab").insert({
            "book_id": book_id,
            "unit_id": unit_id,
            "vocab_item_id": vocab_item_id,
            "surface_word": row["surface_word"],
            "book_meaning": row["meaning"] or None,
            "book_note": row["note"] or None,
            "source_line": f"{source_name or 'excel'}: row {row['row_number']}",
            "item_order": row["row_number"],
        }).execute()
        existing_link_keys.add(link_key)
        created_link_count += 1

    parsed["summary"] = {
        "created_books": created_book_count,
        "created_units": created_unit_count,
        "created_vocab_items": created_vocab_count,
        "updated_vocab_items": updated_vocab_count,
        "created_links": created_link_count,
        "total_rows": len(rows),
    }
    return True, "词汇书导入完成。", parsed
