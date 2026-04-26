from __future__ import annotations

import csv
import json
import random
from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "item_id",
    "module",
    "level",
    "word",
    "primary_meaning_zh",
    "part_of_speech",
    "category",
    "sub_skill",
    "question_type",
    "question_text",
    "correct_answer",
    "wrong_option_1",
    "wrong_option_2",
    "wrong_option_3",
    "explanation",
    "diagnostic_tag",
    "diagnostic_value",
    "difficulty_level",
    "grade_level",
    "frequency_band",
    "source_type",
    "source_note",
    "source_url_primary",
    "source_url_method",
    "is_anchor",
    "is_active",
    "version",
    "sentence",
    "notes_for_codex",
]

SUPPORTED_LEVELS = ("L1", "L2", "L3", "L4", "L5")
DIAGNOSTIC_DATA_SHEET = "diagnostic_vocab_items"
IGNORED_SHEETS = {"README", "tag_dictionary", "Codex_import_prompt"}
DIAGNOSTIC_LEVEL_TARGETS = {
    "L1": 18,
    "L2": 18,
    "L3": 18,
    "L4": 14,
    "L5": 12,
}
ACTIVE_DIAGNOSTIC_VERSIONS = (
    "vocab_diag_v1_l1l2_20260426",
    "vocab_diag_v1_l3l4l5_20260426",
)
QUESTION_TYPES_REQUIRING_SENTENCE = {"polysemy_context"}
QUESTION_TYPES_WITH_OPTION_PREVIEW = {
    "en_to_zh_choice",
    "en_to_zh",
    "zh_to_en",
    "polysemy_context",
    "confusable_choice",
}
UNCERTAIN_OPTION = "\u4e0d\u786e\u5b9a"


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_level(value: Any) -> str:
    return _safe_text(value).upper()


def _parse_bool(value: Any) -> tuple[bool | None, str]:
    text = _safe_text(value).lower()
    if text in {"1", "true", "yes", "y", "是"}:
        return True, ""
    if text in {"0", "false", "no", "n", "否"}:
        return False, ""
    if text == "":
        return None, "empty boolean"
    return None, f"invalid boolean: {value}"


def _select_candidate_sheets(sheet_names: list[str]) -> list[str]:
    if DIAGNOSTIC_DATA_SHEET in sheet_names:
        return [DIAGNOSTIC_DATA_SHEET]
    return [
        sheet_name
        for sheet_name in sheet_names
        if sheet_name not in IGNORED_SHEETS and sheet_name.endswith("_items")
    ]


def _normalize_row(row: dict[str, Any], *, row_number: int, source_name: str, sheet_name: str = "") -> dict:
    parsed_row = {column: _safe_text(row.get(column)) for column in REQUIRED_COLUMNS}
    is_anchor, _anchor_error = _parse_bool(parsed_row["is_anchor"])
    is_active, _active_error = _parse_bool(parsed_row["is_active"])
    parsed_row["is_anchor"] = bool(is_anchor)
    parsed_row["is_active"] = bool(is_active)
    parsed_row["source_name"] = source_name
    parsed_row["sheet_name"] = sheet_name
    parsed_row["row_number"] = row_number
    return parsed_row


def _parse_excel_bytes(file_bytes: bytes, *, source_name: str, sheet_name: str | None = None) -> dict:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    candidate_sheets = (
        [sheet_name]
        if sheet_name and sheet_name in workbook.sheetnames
        else _select_candidate_sheets(workbook.sheetnames)
    )

    if not candidate_sheets:
        return {
            "row_count": 0,
            "rows": [],
            "errors": ["没有找到可导入的首次诊断题库 sheet。"],
            "sheet_names": workbook.sheetnames,
            "selected_sheets": [],
        }

    parsed_rows: list[dict] = []
    errors: list[str] = []
    for candidate_sheet in candidate_sheets:
        worksheet = workbook[candidate_sheet]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            errors.append(f"{candidate_sheet} 是空 sheet。")
            continue

        headers = [_safe_text(value) for value in rows[0]]
        header_index = {header: idx for idx, header in enumerate(headers) if header}
        missing_columns = [column for column in REQUIRED_COLUMNS if column not in header_index]
        if missing_columns:
            errors.append(f"{candidate_sheet} 缺少字段：{', '.join(missing_columns)}")
            continue

        for excel_row_number, row in enumerate(rows[1:], start=2):
            row_dict = {
                column: row[header_index[column]] if header_index[column] < len(row) else ""
                for column in REQUIRED_COLUMNS
            }
            if not any(_safe_text(value) for value in row_dict.values()):
                continue
            parsed_rows.append(
                _normalize_row(
                    row_dict,
                    row_number=excel_row_number,
                    source_name=source_name,
                    sheet_name=candidate_sheet,
                )
            )

    return {
        "row_count": len(parsed_rows),
        "rows": parsed_rows,
        "errors": errors,
        "sheet_names": workbook.sheetnames,
        "selected_sheets": candidate_sheets,
    }


def _parse_csv_bytes(file_bytes: bytes, *, source_name: str) -> dict:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(text.splitlines())
    headers = reader.fieldnames or []
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing_columns:
        return {
            "row_count": 0,
            "rows": [],
            "errors": [f"{source_name} 缺少字段：{', '.join(missing_columns)}"],
            "sheet_names": [source_name],
            "selected_sheets": [source_name],
        }

    parsed_rows = []
    for row_number, row in enumerate(reader, start=2):
        if not any(_safe_text(value) for value in row.values()):
            continue
        parsed_rows.append(
            _normalize_row(row, row_number=row_number, source_name=source_name)
        )

    return {
        "row_count": len(parsed_rows),
        "rows": parsed_rows,
        "errors": [],
        "sheet_names": [source_name],
        "selected_sheets": [source_name],
    }


def _parse_jsonl_bytes(file_bytes: bytes, *, source_name: str) -> dict:
    parsed_rows = []
    errors = []
    for row_number, line in enumerate(file_bytes.decode("utf-8-sig").splitlines(), start=1):
        text = line.strip()
        if not text:
            continue
        try:
            payload = json.loads(text)
        except json.JSONDecodeError as exc:
            errors.append(f"{source_name} 第 {row_number} 行 JSONL 解析失败：{exc}")
            continue
        parsed_rows.append(
            _normalize_row(payload, row_number=row_number, source_name=source_name)
        )
    return {
        "row_count": len(parsed_rows),
        "rows": parsed_rows,
        "errors": errors,
        "sheet_names": [source_name],
        "selected_sheets": [source_name],
    }


def parse_diagnostic_vocab_file(
    file_bytes: bytes,
    *,
    source_name: str,
    sheet_name: str | None = None,
) -> dict:
    suffix = Path(source_name).suffix.lower()
    if suffix == ".xlsx":
        return _parse_excel_bytes(file_bytes, source_name=source_name, sheet_name=sheet_name)
    if suffix == ".csv":
        return _parse_csv_bytes(file_bytes, source_name=source_name)
    if suffix == ".jsonl":
        return _parse_jsonl_bytes(file_bytes, source_name=source_name)
    raise ValueError(f"不支持的诊断题库文件格式：{suffix or source_name}")


def parse_diagnostic_vocab_excel(file_bytes: bytes, *, sheet_name: str | None = None) -> dict:
    return finalize_diagnostic_vocab_parse_result(
        _parse_excel_bytes(file_bytes, source_name="uploaded.xlsx", sheet_name=sheet_name)
    )


def preview_diagnostic_vocab_excel(
    file_bytes: bytes,
    *,
    sheet_name: str | None = None,
    preview_rows: int = 8,
) -> dict:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    selected_sheet = (
        sheet_name
        if sheet_name and sheet_name in workbook.sheetnames
        else _select_candidate_sheets(workbook.sheetnames)[0] if _select_candidate_sheets(workbook.sheetnames) else workbook.sheetnames[0]
    )
    worksheet = workbook[selected_sheet]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [_safe_text(value) for value in rows[0]] if rows else []
    preview = []
    for row in rows[:preview_rows]:
        preview.append([_safe_text(value) for value in row[: len(headers)]])
    return {
        "sheet_names": workbook.sheetnames,
        "active_sheet": selected_sheet,
        "headers": headers,
        "preview_rows": preview,
        "row_count": len(rows),
        "column_count": len(headers),
    }


def merge_diagnostic_parse_results(results: Iterable[dict]) -> dict:
    merged_rows: list[dict] = []
    merged_errors: list[str] = []
    sheet_names: list[str] = []
    selected_sheets: list[str] = []
    for result in results:
        merged_rows.extend(result.get("rows", []))
        merged_errors.extend(result.get("errors", []))
        sheet_names.extend(result.get("sheet_names", []))
        selected_sheets.extend(result.get("selected_sheets", []))

    merged = {
        "row_count": len(merged_rows),
        "rows": merged_rows,
        "errors": merged_errors,
        "sheet_names": sorted(set(sheet_names)),
        "selected_sheets": selected_sheets,
    }
    return finalize_diagnostic_vocab_parse_result(merged)


def validate_diagnostic_vocab_rows(rows: list[dict], *, stop_after: int | None = 200) -> list[str]:
    errors: list[str] = []
    seen_item_ids: set[str] = set()

    for row in rows:
        item_id = _safe_text(row.get("item_id"))
        row_number = row.get("row_number", "?")
        source_name = row.get("source_name") or row.get("sheet_name") or "unknown"
        location = f"{source_name} 第 {row_number} 行"

        def add_error(reason: str) -> None:
            errors.append(f"{location} | item_id={item_id or '<empty>'} | {reason}")

        if not item_id:
            add_error("item_id 不能为空")
        elif item_id in seen_item_ids:
            add_error("item_id 重复")
        else:
            seen_item_ids.add(item_id)

        level = _normalize_level(row.get("level"))
        if not level:
            add_error("level 不能为空")
        elif level not in SUPPORTED_LEVELS:
            add_error(f"level 非法：{level}")

        question_type = _safe_text(row.get("question_type"))
        if not question_type:
            add_error("question_type 不能为空")

        version = _safe_text(row.get("version"))
        if not version:
            add_error("version 不能为空")

        correct_answer = _safe_text(row.get("correct_answer"))
        wrong_options = [
            _safe_text(row.get("wrong_option_1")),
            _safe_text(row.get("wrong_option_2")),
            _safe_text(row.get("wrong_option_3")),
        ]
        if not correct_answer:
            add_error("correct_answer 不能为空")
        if any(not option for option in wrong_options):
            add_error("wrong_option_1~3 不能为空")
        if correct_answer and correct_answer in wrong_options:
            add_error("correct_answer 与 wrong_option 重复")
        non_empty_wrong_options = [option for option in wrong_options if option]
        if len(non_empty_wrong_options) != len(set(non_empty_wrong_options)):
            add_error("wrong_option_1~3 存在重复")

        is_anchor, anchor_error = _parse_bool(row.get("is_anchor"))
        if anchor_error:
            add_error("is_anchor 不是合法布尔值")
        else:
            row["is_anchor"] = bool(is_anchor)

        is_active, active_error = _parse_bool(row.get("is_active"))
        if active_error:
            add_error("is_active 不是合法布尔值")
        else:
            row["is_active"] = bool(is_active)

        if question_type in QUESTION_TYPES_REQUIRING_SENTENCE and not _safe_text(row.get("sentence")):
            add_error(f"{question_type} 必须提供 sentence")

        if stop_after is not None and len(errors) >= stop_after:
            break

    return errors


def summarize_diagnostic_vocab_rows(rows: list[dict], *, validation_errors: list[str] | None = None) -> dict:
    level_counter = Counter()
    question_type_counter = Counter()
    version_counter = Counter()
    anchor_count = 0
    inactive_count = 0
    duplicate_count = 0
    missing_key_field_count = 0
    seen_item_ids: set[str] = set()

    for row in rows:
        item_id = _safe_text(row.get("item_id"))
        if not item_id:
            missing_key_field_count += 1
        elif item_id in seen_item_ids:
            duplicate_count += 1
        else:
            seen_item_ids.add(item_id)

        level_counter[_normalize_level(row.get("level")) or ""] += 1
        question_type_counter[_safe_text(row.get("question_type")) or ""] += 1
        version_counter[_safe_text(row.get("version")) or ""] += 1
        if bool(row.get("is_anchor")):
            anchor_count += 1
        if not bool(row.get("is_active")):
            inactive_count += 1

        key_fields = [
            row.get("item_id"),
            row.get("level"),
            row.get("question_type"),
            row.get("correct_answer"),
            row.get("wrong_option_1"),
            row.get("wrong_option_2"),
            row.get("wrong_option_3"),
            row.get("version"),
        ]
        if any(not _safe_text(value) for value in key_fields):
            missing_key_field_count += 1

    return {
        "total_rows": len(rows),
        "level_counts": dict(level_counter),
        "question_type_counts": dict(question_type_counter),
        "version_counts": dict(version_counter),
        "anchor_count": anchor_count,
        "inactive_count": inactive_count,
        "duplicate_item_id_count": duplicate_count,
        "missing_key_field_count": missing_key_field_count,
        "validation_error_count": len(validation_errors or []),
    }


def finalize_diagnostic_vocab_parse_result(parsed: dict) -> dict:
    rows = parsed.get("rows", [])
    validation_errors = validate_diagnostic_vocab_rows(rows)
    stats = summarize_diagnostic_vocab_rows(rows, validation_errors=validation_errors)
    parsed["errors"] = [*parsed.get("errors", []), *validation_errors]
    parsed["level_counts"] = stats["level_counts"]
    parsed["question_type_counts"] = stats["question_type_counts"]
    parsed["version_counts"] = stats["version_counts"]
    parsed["anchor_count"] = stats["anchor_count"]
    parsed["inactive_count"] = stats["inactive_count"]
    parsed["duplicate_item_id_count"] = stats["duplicate_item_id_count"]
    parsed["missing_key_field_count"] = stats["missing_key_field_count"]
    return parsed


def load_and_parse_diagnostic_vocab_paths(file_paths: list[str]) -> dict:
    parsed_parts = []
    for file_path in file_paths:
        path = Path(file_path)
        with path.open("rb") as file_obj:
            parsed_parts.append(
                parse_diagnostic_vocab_file(
                    file_obj.read(),
                    source_name=path.name,
                )
            )
    return merge_diagnostic_parse_results(parsed_parts)


def build_diagnostic_choice_options(row: dict, *, randomizer: random.Random | None = None) -> list[str]:
    options = [
        _safe_text(row.get("correct_answer")),
        _safe_text(row.get("wrong_option_1")),
        _safe_text(row.get("wrong_option_2")),
        _safe_text(row.get("wrong_option_3")),
        UNCERTAIN_OPTION,
    ]
    deduped = []
    for option in options:
        if option and option not in deduped:
            deduped.append(option)
    if randomizer is None:
        random.shuffle(deduped)
    else:
        randomizer.shuffle(deduped)
    return deduped


def build_diagnostic_question_payload(
    row: dict,
    *,
    randomizer: random.Random | None = None,
) -> dict:
    payload = dict(row)
    payload["options"] = build_diagnostic_choice_options(row, randomizer=randomizer)
    payload["has_uncertain_option"] = UNCERTAIN_OPTION in payload["options"]
    payload["shows_sentence"] = _safe_text(row.get("question_type")) in QUESTION_TYPES_REQUIRING_SENTENCE
    return payload


def select_diagnostic_items_for_test(
    rows: list[dict],
    *,
    target_counts: dict[str, int] | None = None,
    active_versions: Iterable[str] | None = None,
    random_seed: int | None = None,
) -> list[dict]:
    target_counts = target_counts or DIAGNOSTIC_LEVEL_TARGETS
    active_versions = tuple(active_versions or ACTIVE_DIAGNOSTIC_VERSIONS)
    randomizer = random.Random(random_seed)
    selected_rows: list[dict] = []

    for level, target_count in target_counts.items():
        eligible_rows = [
            row
            for row in rows
            if _normalize_level(row.get("level")) == level
            and bool(row.get("is_active"))
            and _safe_text(row.get("version")) in active_versions
        ]
        anchors = [row for row in eligible_rows if bool(row.get("is_anchor"))]
        non_anchors = [row for row in eligible_rows if not bool(row.get("is_anchor"))]

        randomizer.shuffle(anchors)
        randomizer.shuffle(non_anchors)

        picked = anchors[:target_count]
        if len(picked) < target_count:
            needed = target_count - len(picked)
            picked.extend(non_anchors[:needed])

        if len(picked) < target_count:
            raise ValueError(
                f"{level} 可用题目不足：需要 {target_count} 题，实际只有 {len(picked)} 题。"
            )
        selected_rows.extend(picked)

    randomizer.shuffle(selected_rows)
    return selected_rows


def build_diagnostic_preview_summary(rows: list[dict]) -> dict:
    level_counts = Counter(_normalize_level(row.get("level")) for row in rows)
    question_type_counts = Counter(_safe_text(row.get("question_type")) for row in rows)
    return {
        "total_count": len(rows),
        "level_counts": dict(level_counts),
        "question_type_counts": dict(question_type_counts),
    }
